import cv2
import yaml
import os
import numpy as np
from ultralytics import YOLO
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ==================================================
# SCREEN SIZE (DISPLAY ONLY)
# ==================================================
try:
    import tkinter as tk
    _root = tk.Tk()
    _root.withdraw()
    SCREEN_W = _root.winfo_screenwidth() - 100
    SCREEN_H = _root.winfo_screenheight() - 100
    _root.destroy()
except Exception:
    SCREEN_W = 1280
    SCREEN_H = 720

# ==================================================
# CONFIG
# ==================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "cameras.yaml")
BUFFER_FRAMES = 150

# ==================================================
# LOAD YAML
# ==================================================
with open(CONFIG_PATH, "r") as f:
    CONFIG = yaml.safe_load(f)

GLOBAL = CONFIG["global"]
CAMERAS = CONFIG["cameras"]
os.makedirs(GLOBAL["report_dir"], exist_ok=True)

# ==================================================
# CAMERA SELECTION
# ==================================================
def select_cameras_from_yaml(cameras):
    for i, cam in enumerate(cameras):
        print(f"[{i}] {cam['camera_id']} ({cam['source_type']})")

    choice = input("Select camera(s) (0 / 1 / 0,2 / all): ").strip().lower()
    if choice == "all":
        return cameras
    idx = [int(x.strip()) for x in choice.split(",")]
    return [cameras[i] for i in idx]

# ==================================================
# DISPLAY RESIZE
# ==================================================
def resize_frame(frame, max_w, max_h):
    h, w = frame.shape[:2]
    scale = min(max_w / w, max_h / h, 1.0)
    if scale >= 1.0:
        return frame
    return cv2.resize(frame, (int(w * scale), int(h * scale)))

# ==================================================
# ROI LOAD
# ==================================================
def load_line_roi(path, w, h):
    roi_path = path if os.path.isabs(path) else os.path.join(BASE_DIR, path)
    pts = []
    with open(roi_path) as f:
        for line in f:
            if line.strip() and not line.startswith("TYPE"):
                nx, ny = map(float, line.replace(",", " ").split()[:2])
                pts.append((int(nx * w), int(ny * h)))
    if len(pts) != 2:
        raise ValueError("ROI must contain exactly 2 points")
    return pts[0], pts[1]

# ==================================================
# GEOMETRY
# ==================================================
def compute_normal(p1, p2):
    dx = p2[0] - p1[0]
    dy = p2[1] - p1[1]
    n = np.array([-dy, dx], dtype=float)
    n /= np.linalg.norm(n)
    return n

def signed_distance(pt, p1, normal):
    return float(np.dot(np.array(pt) - np.array(p1), normal))

# ==================================================
# REPORTING
# ==================================================
def init_report(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Time", "Camera", "Event", "TrackID"])
        wb.save(path)

def log_event(path, cam, event, tid):
    wb = load_workbook(path)
    ws = wb.active
    assert ws is not None
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        cam,
        event,
        tid
    ])
    wb.save(path)

# ==================================================
# MAIN CAMERA LOOP
# ==================================================
def run_camera(cam_cfg):
    cam_id = cam_cfg["camera_id"]

    if cam_cfg["source_type"] == "file":
        source = cam_cfg["file_path"]
    elif cam_cfg["source_type"] == "rtsp":
        source = input(f"[{cam_id}] Paste RTSP URL: ").strip()
    else:
        source = int(cam_cfg.get("device_index", 0))

    cap = cv2.VideoCapture(source, cv2.CAP_FFMPEG)
    ret, frame = cap.read()
    if not ret:
        print(f"[{cam_id}] Cannot open source")
        return

    h, w = frame.shape[:2]
    p1, p2 = load_line_roi(cam_cfg["roi_file"], w, h)
    normal = compute_normal(p1, p2)

    model = YOLO(GLOBAL["model_path"])

    report_path = os.path.join(GLOBAL["report_dir"], f"footfall_{cam_id}.xlsx")
    init_report(report_path)

    track_state = {}
    in_count = out_count = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        raw = frame.copy()
        results = model.track(
            raw,
            persist=True,
            tracker="bytetrack.yaml",
            classes=[GLOBAL["person_class_id"]],
            verbose=False
        )

        display = resize_frame(raw.copy(), SCREEN_W, SCREEN_H)
        sx = display.shape[1] / raw.shape[1]
        sy = display.shape[0] / raw.shape[0]

        cv2.line(
            display,
            (int(p1[0] * sx), int(p1[1] * sy)),
            (int(p2[0] * sx), int(p2[1] * sy)),
            (255, 0, 0), 2
        )

        if results and results[0].boxes:
            for box in results[0].boxes:
                if box.id is None:
                    continue

                tid = int(box.id[0])
                x1, y1, x2, y2 = map(int, box.xyxy[0])

                # ==================================================
                # HYBRID ANCHOR POINT (INTEGRATED HERE)
                # ==================================================
                h_box = y2 - y1
                cx = (x1 + x2) // 2

                if h_box < 80:                      # far / small person
                    cy = y1 + int(0.5 * h_box)
                    anchor = "C"
                elif h_box > 150:                   # close / large person
                    cy = y1 + int(0.65 * h_box)     # knee-level
                    anchor = "K"
                else:                               # normal case
                    cy = y2                         # feet
                    anchor = "F"

                dist = signed_distance((cx, cy), p1, normal)
                curr_zone = "POS" if dist > 0 else "NEG"

                if tid not in track_state:
                    track_state[tid] = {
                        "last_zone": curr_zone,
                        "cooldown": 0
                    }
                    continue

                state = track_state[tid]
                prev_zone = state["last_zone"]
                event = None

                if state["cooldown"] == 0:
                    if prev_zone == "POS" and curr_zone == "NEG":
                        event = cam_cfg["direction"]["positive_to_negative"]
                    elif prev_zone == "NEG" and curr_zone == "POS":
                        event = cam_cfg["direction"]["negative_to_positive"]

                if event:
                    if event.upper() == "IN":
                        in_count += 1
                    elif event.upper() == "OUT":
                        out_count += 1

                    log_event(report_path, cam_id, event, tid)
                    state["cooldown"] = BUFFER_FRAMES

                if state["cooldown"] > 0:
                    state["cooldown"] -= 1

                state["last_zone"] = curr_zone

                # ==================================================
                # DRAW BB + DEBUG INFO
                # ==================================================
                x1d, y1d = int(x1 * sx), int(y1 * sy)
                x2d, y2d = int(x2 * sx), int(y2 * sy)

                color = (0, 255, 0) if state["cooldown"] > 0 else (0, 255, 255)
                cv2.rectangle(display, (x1d, y1d), (x2d, y2d), color, 2)
                cv2.circle(display, (int(cx * sx), int(cy * sy)), 4, (0, 0, 255), -1)

                label = (
                    f"ID:{tid} "
                    f"Z:{curr_zone} "
                    f"P:{prev_zone} "
                    f"D:{dist:.1f} "
                    f"A:{anchor} "
                    f"CD:{state['cooldown']}"
                )

                cv2.putText(
                    display,
                    label,
                    (x1d, y1d - 6),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.45,
                    color,
                    2
                )

        cv2.putText(display, f"IN: {in_count}", (20, 40),
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 3)
        cv2.putText(display, f"OUT: {out_count}", (20, 90),
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)

        cv2.imshow(cam_id, display)
        if cv2.waitKey(1) & 0xFF == 27:
            break

    cap.release()
    cv2.destroyAllWindows()

# ==================================================
# ENTRY POINT
# ==================================================
if __name__ == "__main__":
    cams = select_cameras_from_yaml(CAMERAS)
    for cam in cams:
        run_camera(cam)
